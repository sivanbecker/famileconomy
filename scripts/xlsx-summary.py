#!/usr/bin/env python3
"""
Summarise monthly expenses from MAX or CAL XLSX files.

Reads XLSX files and prints total expenses per month (and optionally per card).
Amounts are read as-is from the XLSX (ILS only) and output in ₪.

Usage:
  # Single file
  python xlsx-summary.py --provider MAX --file /path/to/file.xlsx
  python xlsx-summary.py --provider CAL --file /path/to/file.xlsx --per-card

  # Folder (all XLSX files in folder)
  python xlsx-summary.py --provider MAX --folder /path/to/folder
  python xlsx-summary.py --provider MAX --folder /path/to/folder --per-card

  # Use configured xlsx_base from providers.json
  python xlsx-summary.py --provider MAX --year 2026
  python xlsx-summary.py --provider MAX --year 2026 --per-card
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


# ─── Config ───────────────────────────────────────────────────────────────────


def load_config(config_path: Path | None = None) -> dict:
    if config_path is None:
        script_dir = Path(__file__).parent
        config_path = script_dir.parent / "config" / "providers.json"
    if not config_path.exists():
        raise FileNotFoundError(f"Config file not found: {config_path}")
    with open(config_path) as f:
        config = json.load(f)
    root = config.get("google_drive_root", "")
    for provider in config.values():
        if not isinstance(provider, dict):
            continue
        for key in ("xlsx_base", "csv_base"):
            if key in provider and root:
                provider[key] = str(Path(root) / provider[key])
    return config


# ─── Month/year extraction (shared with other scripts) ────────────────────────


def extract_month_year(cell_value: str) -> tuple[int, int]:
    """Returns (month, year). Handles MM/YYYY (MAX) and DD/MM/YYYY embedded (CAL)."""
    if not cell_value or not isinstance(cell_value, str):
        raise ValueError(f"Invalid cell value: {cell_value!r}")
    cell_value = cell_value.strip()
    match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", cell_value)
    if match:
        day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        if 1 <= month <= 12 and 2020 <= year <= 2100:
            return month, year
    parts = cell_value.split("/")
    if len(parts) == 2:
        try:
            month, year = int(parts[0]), int(parts[1])
            if 1 <= month <= 12 and 2020 <= year <= 2100:
                return month, year
        except ValueError:
            pass
    raise ValueError(f"Cannot parse month/year from: {cell_value!r}")


# ─── Amount parsing ───────────────────────────────────────────────────────────


def parse_amount_ils(raw) -> float | None:
    """Parse a cell value as an ILS amount. Returns None if not a number."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).replace(",", "").replace("₪", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ─── MAX parsing ──────────────────────────────────────────────────────────────

# MAX column layout (0-based):
#   0: transaction date, 1: description, 2: category
#   3: card last-4, 4: ?, 5: amount (ILS), ...
_MAX_COL_CARD = 3   # "4 ספרות אחרונות של כרטיס האשראי"
_MAX_COL_AMOUNT = 5  # "סכום העסקה בש\"ח"
_MAX_HEADER_MARKER = "תאריך עסקה"


def _find_max_header_row(ws) -> int | None:
    """Return 1-based row index of the header row, or None."""
    for i in range(1, min(ws.max_row, 20) + 1):
        cell = ws.cell(row=i, column=1).value
        if cell and str(cell).startswith(_MAX_HEADER_MARKER):
            return i
    return None


def summarise_max(wb: openpyxl.Workbook) -> list[dict]:
    """
    Returns a list of dicts: {month, year, card, amount_ils}
    for every data row in the MAX workbook.
    """
    rows = []
    ws = wb.worksheets[0]

    # Extract month/year from A3
    a3 = wb.active["A3"].value
    if not a3:
        raise ValueError("MAX XLSX: A3 is empty, cannot read month/year")
    month, year = extract_month_year(str(a3))

    header_row = _find_max_header_row(ws)
    if header_row is None:
        raise ValueError("MAX XLSX: header row not found")

    for row in ws.iter_rows(min_row=header_row + 1):
        date_cell = row[0].value
        if not date_cell:
            continue
        card_val = row[_MAX_COL_CARD].value
        card = str(card_val).strip() if card_val else None
        if not card or len(card) != 4 or not card.isdigit():
            continue
        amount_raw = row[_MAX_COL_AMOUNT].value
        amount = parse_amount_ils(amount_raw)
        if amount is None:
            continue
        rows.append({"month": month, "year": year, "card": card, "amount_ils": abs(amount)})
    return rows


# ─── CAL parsing ──────────────────────────────────────────────────────────────

# CAL column layout (0-based):
#   0: transaction date, 1: description, 2: tx amount, 3: charge amount, ...
_CAL_COL_CHARGE_AMOUNT = 3  # "סכום חיוב"
_CAL_COL_TX_AMOUNT = 2      # fallback: "סכום עסקה"
_CAL_CARD_RE = re.compile(r"המסתיים\s+ב-(\d{4})")
_CAL_CHARGE_DATE_RE = re.compile(r"עסקאות לחיוב ב-(\d{2})/(\d{2})/(\d{4})")
_CAL_HEADER_MARKER = "תאריך"


def _find_cal_header_row(ws) -> int | None:
    for i in range(1, min(ws.max_row, 20) + 1):
        cell = ws.cell(row=i, column=1).value
        if cell:
            first_line = str(cell).split("\n")[0].strip()
            if first_line.startswith(_CAL_HEADER_MARKER):
                return i
    return None


def _parse_cal_amount(raw) -> float | None:
    """CAL amounts are formatted like '₪1,234.56' or plain numbers."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).replace(",", "").replace("₪", "").strip()
    if s.startswith("$") or s.startswith("€") or s.startswith("£"):
        return None  # skip foreign currency rows for ILS totals
    try:
        return float(s)
    except ValueError:
        return None


def summarise_cal(wb: openpyxl.Workbook) -> list[dict]:
    ws = wb.worksheets[0]

    # Card from A1
    a1 = str(ws.cell(row=1, column=1).value or "")
    card_match = _CAL_CARD_RE.search(a1)
    if not card_match:
        raise ValueError("CAL XLSX: card identifier not found in A1")
    card = card_match.group(1)

    # Month/year from charge-date line (search first 10 rows)
    month = year = None
    for i in range(1, min(ws.max_row, 10) + 1):
        cell_val = str(ws.cell(row=i, column=1).value or "")
        m = _CAL_CHARGE_DATE_RE.search(cell_val)
        if m:
            month, year = int(m.group(2)), int(m.group(3))
            break
    if month is None:
        # Fallback: A3
        a3 = wb.active["A3"].value
        if a3:
            month, year = extract_month_year(str(a3))
    if month is None:
        raise ValueError("CAL XLSX: could not determine charge month/year")

    header_row = _find_cal_header_row(ws)
    if header_row is None:
        raise ValueError("CAL XLSX: header row not found")

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1):
        date_cell = row[0].value
        if not date_cell:
            continue
        # Try charge amount first, fall back to tx amount
        charge_raw = row[_CAL_COL_CHARGE_AMOUNT].value
        amount = _parse_cal_amount(charge_raw)
        if amount is None:
            tx_raw = row[_CAL_COL_TX_AMOUNT].value
            amount = _parse_cal_amount(tx_raw)
        if amount is None:
            continue
        # Skip pending rows (notes column 6, 0-based)
        notes_raw = row[6].value if len(row) > 6 else None
        notes = str(notes_raw or "").strip()
        if "עסקה בקליטה" in notes:
            continue
        rows.append({"month": month, "year": year, "card": card, "amount_ils": abs(amount)})
    return rows


# ─── Aggregation & display ────────────────────────────────────────────────────


def aggregate(rows: list[dict], per_card: bool) -> dict:
    """
    Returns an ordered dict keyed by (year, month) or (year, month, card).
    Values are total ILS amounts (floats).
    """
    totals: dict[tuple, float] = defaultdict(float)
    for r in rows:
        if per_card:
            key = (r["year"], r["month"], r["card"])
        else:
            key = (r["year"], r["month"])
        totals[key] += r["amount_ils"]
    return dict(sorted(totals.items()))


def format_totals(provider: str, totals: dict, per_card: bool) -> list[str]:
    lines = []
    for key, total in totals.items():
        year, month = key[0], key[1]
        month_name = MONTH_NAMES[month - 1]
        ils = round(total)
        if per_card:
            card = key[2]
            lines.append(f"{provider} {card} {month_name} {year}: {ils:,} ₪")
        else:
            lines.append(f"{provider} {month_name} {year}: {ils:,} ₪")
    return lines


# ─── File processing ──────────────────────────────────────────────────────────


def process_file(file_path: Path, provider: str) -> list[dict]:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        raise ValueError(f"Failed to load {file_path.name}: {e}")
    if provider == "MAX":
        return summarise_max(wb)
    elif provider == "CAL":
        return summarise_cal(wb)
    else:
        raise ValueError(f"Unknown provider: {provider}")


# ─── Public API for runner.py ─────────────────────────────────────────────────


def summarise_single(
    file_path: str,
    provider: str,
    year: int | None = None,  # unused, kept for consistent runner signature
    per_card: bool = False,
) -> tuple[bool, str]:
    try:
        rows = process_file(Path(file_path), provider.upper())
        totals = aggregate(rows, per_card)
        lines = format_totals(provider.upper(), totals, per_card)
        return (True, "\n".join(lines) if lines else "(no expenses found)")
    except Exception as e:
        return (False, f"✗ {Path(file_path).name}: {e}")


def summarise_batch(
    folder_path: str | None,
    provider: str,
    year: int | None = None,
    per_card: bool = False,
) -> list[tuple[bool, str]]:
    provider = provider.upper()

    if folder_path is None:
        # Resolve from config
        try:
            config = load_config()
        except FileNotFoundError as e:
            return [(False, str(e))]
        xlsx_base = config.get(provider, {}).get("xlsx_base")
        if not xlsx_base:
            return [(False, f"No xlsx_base configured for {provider} in providers.json")]
        folder = Path(xlsx_base)
    else:
        folder = Path(folder_path)

    if year is not None:
        folder = folder / str(year)

    xlsx_files = sorted(folder.glob("*.xlsx"))
    if not xlsx_files:
        # Try one level of subfolders (year dirs)
        xlsx_files = sorted(folder.glob("*/*.xlsx"))
    if not xlsx_files:
        return [(False, f"No XLSX files found under: {folder}")]

    all_rows: list[dict] = []
    errors: list[tuple[bool, str]] = []
    for f in xlsx_files:
        try:
            all_rows.extend(process_file(f, provider))
        except Exception as e:
            errors.append((False, f"✗ {f.name}: {e}"))

    if not all_rows and not errors:
        return [(False, "No data found")]

    totals = aggregate(all_rows, per_card)
    lines = format_totals(provider, totals, per_card)
    results = [(True, line) for line in lines]
    return results + errors


# ─── CLI ──────────────────────────────────────────────────────────────────────


def _xlsx_base(config: dict, provider: str) -> Path:
    xlsx_base = config.get(provider, {}).get("xlsx_base")
    if not xlsx_base:
        raise SystemExit(f"Error: no xlsx_base configured for {provider} in providers.json")
    return Path(xlsx_base)


def main():
    try:
        config = load_config()
    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

    parser = argparse.ArgumentParser(
        description="Summarise monthly expenses from MAX or CAL XLSX files"
    )
    parser.add_argument("--provider", required=True, choices=["MAX", "CAL"])
    parser.add_argument("--file", help="Path to a single XLSX file")
    parser.add_argument("--folder", help="Folder of XLSX files (defaults to xlsx_base from config)")
    parser.add_argument(
        "--year",
        type=int,
        help="Restrict to a year subfolder inside xlsx_base (e.g. 2026)",
    )
    parser.add_argument(
        "--per-card",
        action="store_true",
        help="Break down totals per card instead of per month total",
    )
    args = parser.parse_args()

    if args.file and (args.folder or args.year):
        print("Error: --file is mutually exclusive with --folder / --year", file=sys.stderr)
        sys.exit(1)

    provider = args.provider.upper()

    # ── Single file ────────────────────────────────────────────────────────
    if args.file:
        ok, msg = summarise_single(args.file, provider, per_card=args.per_card)
        print(msg)
        sys.exit(0 if ok else 1)

    # ── Folder / year mode — resolve folder from config when not explicit ──
    if args.folder:
        folder = Path(args.folder)
    else:
        folder = _xlsx_base(config, provider)

    if args.year:
        folder = folder / str(args.year)

    if not folder.exists():
        print(f"Error: folder not found: {folder}", file=sys.stderr)
        sys.exit(1)

    # Collect XLSX files: either directly in folder or recursively in year subfolders
    xlsx_files = sorted(folder.glob("*.xlsx"))
    if not xlsx_files:
        # No files at this level — try one level of subfolders (year dirs)
        xlsx_files = sorted(folder.glob("*/*.xlsx"))

    if not xlsx_files:
        print(f"No XLSX files found under: {folder}", file=sys.stderr)
        sys.exit(1)

    all_rows: list[dict] = []
    errors: list[str] = []
    for f in xlsx_files:
        try:
            all_rows.extend(process_file(f, provider))
        except Exception as e:
            errors.append(f"✗ {f.name}: {e}")

    for err in errors:
        print(err, file=sys.stderr)

    totals = aggregate(all_rows, args.per_card)
    for line in format_totals(provider, totals, args.per_card):
        print(line)

    sys.exit(0 if not errors else 1)


if __name__ == "__main__":
    main()
