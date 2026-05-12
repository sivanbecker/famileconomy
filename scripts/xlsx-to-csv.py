#!/usr/bin/env python3
"""
Convert MAX or CAL XLSX files to CSV format.

Detects provider, extracts month/year and card identifiers, validates the file structure,
renames the input file, and outputs CSV files with standardized naming.

Usage:
  python xlsx-to-csv.py --input /path/to/input --output /path/to/output --provider MAX

  Environment variables (overridden by arguments):
    INPUT_FOLDER: path to input folder
    OUTPUT_FOLDER: path to output folder
    PROVIDER: MAX or CAL
"""

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter


def load_config(config_path: Path = None) -> dict:
    """Load provider configuration from JSON file."""
    if config_path is None:
        # Look for config relative to script location
        script_dir = Path(__file__).parent
        config_path = script_dir.parent / "config" / "providers.json"

    if not config_path.exists():
        raise FileNotFoundError(f"Config file not found: {config_path}")

    with open(config_path, "r") as f:
        return json.load(f)


def parse_arguments(config: dict) -> argparse.Namespace:
    """Parse CLI arguments, environment variables, and config file."""
    parser = argparse.ArgumentParser(
        description="Convert MAX or CAL XLSX to CSV with standardized naming"
    )
    parser.add_argument(
        "--input",
        default=os.getenv("INPUT_FOLDER"),
        help="Input folder containing XLSX files (or INPUT_FOLDER env var, or config)",
    )
    parser.add_argument(
        "--output",
        default=os.getenv("OUTPUT_FOLDER"),
        help="Output folder for CSV files (or OUTPUT_FOLDER env var, or config)",
    )
    parser.add_argument(
        "--provider",
        default=os.getenv("PROVIDER"),
        choices=["MAX", "CAL"],
        help="Provider type: MAX or CAL (or PROVIDER env var)",
    )
    parser.add_argument(
        "--file",
        help="XLSX filename to process (relative to input folder). Omit for batch mode",
    )
    parser.add_argument(
        "--year",
        type=int,
        help="Year folder (default: detect from XLSX). Required for batch mode",
    )
    parser.add_argument(
        "--batch",
        action="store_true",
        help="Process all XLSX files in the year folder",
    )

    args = parser.parse_args()

    # Validate provider first
    if not args.provider:
        print("Error: --provider or PROVIDER env var required", file=sys.stderr)
        sys.exit(1)

    # Validate batch vs single file
    if args.batch and args.file:
        print("Error: --batch and --file are mutually exclusive", file=sys.stderr)
        sys.exit(1)
    if not args.batch and not args.file:
        print("Error: provide either --file or --batch", file=sys.stderr)
        sys.exit(1)
    if args.batch and not args.year:
        print("Error: --year required for batch mode", file=sys.stderr)
        sys.exit(1)

    # Merge with config defaults (CLI args take precedence)
    provider_config = config.get(args.provider, {})

    if not args.input:
        args.input = provider_config.get("xlsx_base")
    if not args.output:
        args.output = provider_config.get("csv_base")

    if not args.input:
        print(
            f"Error: --input, INPUT_FOLDER env var, or config[{args.provider}].xlsx_base required",
            file=sys.stderr,
        )
        sys.exit(1)
    if not args.output:
        print(
            f"Error: --output, OUTPUT_FOLDER env var, or config[{args.provider}].csv_base required",
            file=sys.stderr,
        )
        sys.exit(1)

    # Store provider config for later validation
    args.provider_config = provider_config

    return args


def extract_month_year(cell_value: str) -> tuple[int, int]:
    """
    Extract month and year from cell value.
    Supports formats: "MM/YYYY" (MAX) and "DD/MM/YYYY" (CAL, possibly embedded in text).
    Returns: (month, year) as integers.
    """
    import re

    if not cell_value or not isinstance(cell_value, str):
        raise ValueError(f"Invalid month/year cell value: {cell_value}")

    cell_value = cell_value.strip()

    # Try to find DD/MM/YYYY pattern (CAL format, may be embedded in text)
    match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", cell_value)
    if match:
        day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        if not (1 <= month <= 12):
            raise ValueError(f"Invalid month: {month}")
        if not (2020 <= year <= 2100):
            raise ValueError(f"Invalid year: {year}")
        return month, year

    # Try MM/YYYY format (MAX format)
    parts = cell_value.split("/")
    if len(parts) == 2:
        try:
            month, year = int(parts[0]), int(parts[1])
            if not (1 <= month <= 12):
                raise ValueError(f"Invalid month: {month}")
            if not (2020 <= year <= 2100):
                raise ValueError(f"Invalid year: {year}")
            return month, year
        except (ValueError, IndexError):
            pass

    raise ValueError(f"Cannot parse month/year from: {cell_value}")


def get_month_name(month: int) -> str:
    """Convert month number to English name."""
    months = [
        "january",
        "february",
        "march",
        "april",
        "may",
        "june",
        "july",
        "august",
        "september",
        "october",
        "november",
        "december",
    ]
    return months[month - 1]


def extract_max_cards(wb: openpyxl.Workbook, valid_cards: list[str] = None) -> list[str]:
    """
    Extract unique card last-4 digits from MAX XLSX.
    Card column header: "4 ספרות אחרונות של כרטיס האשראי"
    Searches first 10 rows for the header (not always in row 1).
    Validates against valid_cards list if provided.
    """
    cards = set()

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Find the column with the card header (search first 10 rows)
        card_col = None
        header_row = None
        for row_idx in range(1, 11):
            for col in ws.iter_cols(min_row=row_idx, max_row=row_idx):
                for cell in col:
                    if cell.value and "4 ספרות אחרונות של כרטיס האשראי" in str(
                        cell.value
                    ):
                        card_col = cell.column
                        header_row = row_idx
                        break
                if card_col:
                    break
            if card_col:
                break

        if not card_col:
            raise ValueError(
                f"Sheet '{sheet}': Could not find card column '4 ספרות אחרונות של כרטיס האשראי' in first 10 rows"
            )

        # Extract card numbers from data rows (starting after header)
        for row in ws.iter_rows(min_row=header_row + 1):
            cell_value = row[card_col - 1].value
            if cell_value and isinstance(cell_value, str) and len(cell_value.strip()) == 4:
                cards.add(cell_value.strip())

    if not cards:
        raise ValueError("No card numbers found in MAX XLSX")

    # Validate against expected cards
    cards_list = list(cards)
    if valid_cards:
        for card in cards_list:
            if card not in valid_cards:
                raise ValueError(
                    f"Card {card} not in valid_cards {valid_cards} for MAX. File structure may have changed or file was in wrong folder."
                )

    return cards_list


def extract_cal_card(wb: openpyxl.Workbook, valid_cards: list[str] = None) -> str:
    """
    Extract card last-4 digit from CAL XLSX.
    CAL has only one sheet. Row 1, Col A contains text like:
    "פירוט עסקאות לחשבון מזרחי-טפחות 123-123456 לכרטיס ויזה זהב עסקי המסתיים ב-4321"
    Extract the 4-digit number after "המסתיים ב-"
    Validates against valid_cards list if provided.
    """
    ws = wb.active
    cell_value = str(ws["A1"].value or "")

    if "המסתיים ב-" not in cell_value:
        raise ValueError(
            "CAL XLSX: Could not find 'המסתיים ב-' in cell A1. Structure may have changed."
        )

    # Extract 4 digits after "המסתיים ב-"
    import re

    match = re.search(r"המסתיים ב-(\d{4})", cell_value)
    if not match:
        raise ValueError(f"CAL XLSX: Could not parse card number from A1: {cell_value}")

    card = match.group(1)

    # Validate against expected cards
    if valid_cards and card not in valid_cards:
        raise ValueError(
            f"Card {card} not in valid_cards {valid_cards} for CAL. File structure may have changed or file was in wrong folder."
        )

    return card


def process_file(
    input_path: Path, provider: str, valid_cards: list[str] = None
) -> tuple[str, str, list[str]]:
    """
    Load XLSX, validate structure, extract metadata.
    Returns: (month_name, year_str, card_list)
    """
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
    except Exception as e:
        raise ValueError(f"Failed to load XLSX: {e}")

    # Extract month/year from row 3, col A
    cell_value = wb.active["A3"].value
    if not cell_value:
        raise ValueError("Row 3, Col A is empty. Cannot extract month/year.")

    month, year = extract_month_year(str(cell_value))
    month_name = get_month_name(month)
    year_str = str(year)

    # Extract and validate cards
    if provider == "MAX":
        cards = extract_max_cards(wb, valid_cards)
    elif provider == "CAL":
        card = extract_cal_card(wb, valid_cards)
        cards = [card]
    else:
        raise ValueError(f"Unknown provider: {provider}")

    wb.close()
    return month_name, year_str, cards


def rename_input_file(input_path: Path, provider: str, month: str, year: str, cards: list[str]) -> Path:
    """Rename input XLSX file with standardized naming."""
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    cards_str = "-".join(cards)
    new_name = f"{provider.lower()}-{month}-{year}-{timestamp}-{cards_str}.xlsx"
    new_path = input_path.parent / new_name

    input_path.rename(new_path)
    return new_path


def convert_sheet_to_csv(
    input_path: Path, output_dir: Path, provider: str, month: str, year: str, cards: list[str], sheet_index: int = 0
) -> str:
    """
    Convert a sheet from XLSX to CSV.
    For MAX: sheet_index 0 = regular, 1 = foreign. Output: ...-1.csv, ...-2.csv
    For CAL: sheet_index 0 only. Output: ....csv
    """
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    cards_str = "-".join(cards)

    if provider == "MAX":
        suffix = f"-{sheet_index + 1}.csv"
    else:
        suffix = ".csv"

    output_filename = f"{provider.lower()}-{month}-{year}-{timestamp}-{cards_str}{suffix}"
    output_path = output_dir / output_filename

    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.worksheets[sheet_index]

    with open(output_path, "w", encoding="utf-8-sig", newline="") as csvfile:
        import csv

        writer = csv.writer(csvfile)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)

    wb.close()
    return str(output_path)


def process_single_file(
    input_path: Path,
    output_base: Path,
    provider: str,
    valid_cards: list[str],
    year_override: int = None,
) -> tuple[bool, str]:
    """
    Process a single XLSX file.
    Returns: (success: bool, message: str)
    """
    try:
        # Process XLSX to extract month/year and cards
        month_name, year_str, cards = process_file(input_path, provider, valid_cards)

        # Validate year against override if provided
        if year_override:
            if int(year_str) != year_override:
                return (
                    False,
                    f"Year mismatch: XLSX contains {year_str} but expected {year_override}",
                )

        final_year = year_str

        # Construct output directory with year and month
        output_dir = output_base / final_year / month_name
        output_dir.mkdir(parents=True, exist_ok=True)

        # Rename input file
        renamed_input = rename_input_file(input_path, provider, month_name, year_str, cards)

        # Convert sheets to CSV
        wb = openpyxl.load_workbook(renamed_input, data_only=True)
        sheet_count = len(wb.sheetnames)
        wb.close()

        output_files = []
        for sheet_idx in range(sheet_count):
            csv_path = convert_sheet_to_csv(
                renamed_input, output_dir, provider, month_name, year_str, cards, sheet_idx
            )
            output_files.append(csv_path)

        return (
            True,
            f"✓ {input_path.name}: {len(output_files)} CSV(s) → {month_name}/{year_str}",
        )

    except Exception as e:
        return (False, f"✗ {input_path.name}: {e}")


def main():
    try:
        # Load config first
        config = load_config()
    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

    args = parse_arguments(config)

    provider = args.provider.upper()
    valid_cards = args.provider_config.get("valid_cards", [])
    input_base = Path(args.input).expanduser().resolve()
    output_base = Path(args.output).expanduser().resolve()

    # Batch mode
    if args.batch:
        year_folder = input_base / str(args.year)
        if not year_folder.exists():
            print(f"Error: Year folder does not exist: {year_folder}", file=sys.stderr)
            sys.exit(1)

        # Find all XLSX files
        xlsx_files = sorted(year_folder.glob("*.xlsx"))
        if not xlsx_files:
            print(f"No XLSX files found in {year_folder}", file=sys.stderr)
            sys.exit(1)

        print(f"Processing {len(xlsx_files)} file(s) from {year_folder.name}/")
        results = []
        for xlsx_file in xlsx_files:
            success, message = process_single_file(
                xlsx_file, output_base, provider, valid_cards, args.year
            )
            results.append((success, message))
            print(message)

        # Summary
        successful = sum(1 for s, _ in results if s)
        print(f"\n✓ Processed {successful}/{len(xlsx_files)} files")
        return 0 if successful == len(xlsx_files) else 1

    # Single file mode
    else:
        filename = args.file
        if args.year:
            input_path = input_base / str(args.year) / filename
        else:
            input_path = input_base / filename

        if not input_path.exists():
            print(f"Error: Input file does not exist: {input_path}", file=sys.stderr)
            sys.exit(1)

        success, message = process_single_file(
            input_path, output_base, provider, valid_cards, args.year
        )
        print(message)
        return 0 if success else 1


if __name__ == "__main__":
    main()
